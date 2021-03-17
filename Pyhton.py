import pandas as pd
from openpyxl import load_workbook

sheet_1 = pd.read_excel('Book1.xlsx', sheet_name = 0)
sheet_2 = pd.read_excel('Book1.xlsx', sheet_name = 1)
sheet_3 = pd.read_excel('Book1.xlsx', sheet_name = 2)
sheet_4 = pd.read_excel('Book1.xlsx', sheet_name = 3)
sheet_5 = pd.read_excel('Book1.xlsx', sheet_name = 4)

all_data_1 = pd.merge(sheet_1, sheet_2, how='left')
all_data_2 = pd.merge(all_data_1, sheet_3, how='left')
all_data_3 = pd.merge(all_data_2, sheet_4, how='left')
all_data_4 = pd.merge(all_data_3, sheet_5, how='left')


print(all_data_4)
# df = pd.read_excel('Book1.xlsx', sheet_name = 'mastersheet')
psnum = int(input("Enter the PS Number : "))

listOfColumns = ["PS number","Display Name","Colour","Height","School","Sex","Age","Famsize","Mjob","Fjob",
                 "Car","Room1","Reason","Guardian","Traveltime","Studytime","Failures","Schoolsup",
                 "Group","Date","Activities","Nursery","Higher","Internet","Romantic","Absences",
                 "Class","Weight","SchoolName","Address","City","Postal","Phone","Principal","Community","Tstreet"
                 ,"email","School Type","Country","Admission","Date","Year"]

df1 = pd.DataFrame(all_data_4, columns = listOfColumns)
df1.set_index("PS number",inplace = True)

result = df1.loc[psnum]

path = r"Book1.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
if 'mastersheet' in book.sheetnames:
    pfd = book['mastersheet'] 
    book.remove(pfd)
result.to_excel(writer, sheet_name = 'mastersheet')

writer.save()
writer.close()




