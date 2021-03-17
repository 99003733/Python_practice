
import pandas as pd
from openpyxl import load_workbook

def count_and_display_sheet_names():
    workbook = load_workbook(filename="Book1.xlsx")
    res = workbook.sheetnames
    print(res)
    print(len(res))
    
count_and_display_sheet_names()

def sheets_access():
    workbook = load_workbook(filename="Book1.xlsx")
    res = workbook.sheetnames
    print("Total number of sheets in excel file are " +  str(len(res)))
    print("Do you want to search data in all sheets: ")
    answer = input()
    sheets = {}
    if(answer.lower() == 'yes'):
        for i in range(len(res)):
            sheets["Sheet" + str(i)] = pd.read_excel("Book1.xlsx",engine = "openpyxl",sheet_name = i)
#             print(sheets["Sheet" + str(i)])
    else:
        print("Enter the number of sheets you want to scan")
        num = int(input())
        for i in range(num):
            sheets["Sheet" + str(i)] = pd.read_excel("Book1.xlsx",engine = "openpyxl",sheet_name = i)
#             print(sheets["Sheet" + str(i)])
    return sheets
            


def search_id():
    sheeter = sheets_access()
    ps_number = int(input())
    for df in sheeter.values():
        result = df[df['PS number'] == ps_number]
        print(result)
        
    
#         path = r"Book1.xlsx"
#         book = load_workbook(path)
#         writer = pd.ExcelWriter(path, engine = 'openpyxl')
#         writer.book = book
#         if 'mastersheet' in book.sheetnames:
#             pfd = book['mastersheet'] 
#             book.remove(pfd)
#         result.to_excel(writer, sheet_name = 'mastersheet') 
        
#     writer.save()
#     writer.close()
    
search_id()
            
    
