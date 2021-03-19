import os

def find_files():
    print("Enter the file name : ")
    filename = input()
    result = []
    for root, dir, files in os.walk("D:"):
        if filename in files:
            result.append(os.path.join(root, filename))
    return result

def file_finder():
    searched_data=find_files()
    #save_data_mastersheet(searched_data)
    print('Wanna search more?????...')
    ans_more=input()
    while ans_more.lower() == 'yes':
        more_searched_data = []
        more_searched_data=find_files()
        searched_data.extend(more_searched_data)
        print('Wanna search more?????...')
        ans_more=input()
        if ans_more.lower()=='no':
            break

    return searched_data

import pandas as pd
from openpyxl import load_workbook

def sheets_access():
    hmm = file_finder()
    for j in range(len(hmm)):
        workbook = load_workbook(filename=hmm[j])
        res = workbook.sheetnames
        print(res)
        print("Total number of sheets in excel file are " + str(len(res)))
        sheets = {}                                          # empty dictionary
        for i in range(len(res)):                        # for loop for multiple sheets in the excelbook
            sheets["Sheet" + str(i)] = pd.read_excel(hmm[j],engine = "openpyxl", sheet_name=i)
            sheets["Sheet" + str(i)].dropna(axis=1, how='all', inplace=True)
#             print(sheets["Sheet" + str(i)])
    
    return sheets


def search_id(df, ps_number):
    result = df[df['nschid'] == ps_number]
    return result                                     # result is the dataframe of all the matching information of the unique ID

"""
Function to match unique ID in sheets present in the excel book
"""
def match_unique():
    dict_sheets = sheets_access()                        # the function sheet access is assigned to the new variable
    results={}
    print("Enter the Unique ID you want to search : ")
    num = int(input())
    for j in range(len(dict_sheets.keys())):
        results['from_sheet'+str(j)]=search_id(dict_sheets['Sheet'+str(j)],num)

    if len(dict_sheets.keys()) > 1:
        all_data = pd.merge(results['from_sheet0'], results['from_sheet1'], how='left')
        for j in range(2, len(dict_sheets.keys())):
            all_data = pd.merge(all_data, results['from_sheet' + str(j)], how='left')
    else:
        all_data = results['from_sheet0']

    return all_data

match_unique()
